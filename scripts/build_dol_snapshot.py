#!/usr/bin/env python3

import argparse
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


CERTIFIED_STATUSES = {"Certified"}


def normalized_name(value):
    return re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()


def load_entity_map(path):
    raw = json.loads(path.read_text())
    return {
        company_id: {normalized_name(entity) for entity in entities}
        for company_id, entities in raw.items()
    }


def summarize_workbook(path, employer_column, status_column, entity_map):
    reverse_index = {
        entity: company_id
        for company_id, entities in entity_map.items()
        for entity in entities
    }
    totals = Counter()
    certified = Counter()
    status_counts = {company_id: Counter() for company_id in entity_map}
    matched_entities = {company_id: set() for company_id in entity_map}

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    employer_index = header.index(employer_column)
    status_index = header.index(status_column)

    for row in rows:
        entity = normalized_name(row[employer_index])
        company_id = reverse_index.get(entity)
        if not company_id:
            continue
        status = str(row[status_index] or "").strip()
        totals[company_id] += 1
        status_counts[company_id][status] += 1
        matched_entities[company_id].add(str(row[employer_index]).strip())
        if status in CERTIFIED_STATUSES:
            certified[company_id] += 1

    workbook.close()
    return totals, certified, status_counts, matched_entities


def main():
    parser = argparse.ArgumentParser(
        description="Build a reproducible FY2025 DOL LCA/PERM employer snapshot."
    )
    parser.add_argument("--lca", required=True, type=Path)
    parser.add_argument("--perm", required=True, type=Path)
    parser.add_argument(
        "--entities",
        type=Path,
        default=Path("data/employer-entities.json"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/dol-fy2025-snapshot.json"),
    )
    parser.add_argument(
        "--javascript-output",
        type=Path,
        default=Path("data/dol-fy2025-snapshot.js"),
    )
    args = parser.parse_args()

    entity_map = load_entity_map(args.entities)
    lca_total, lca_certified, lca_statuses, lca_entities = summarize_workbook(
        args.lca, "EMPLOYER_NAME", "CASE_STATUS", entity_map
    )
    perm_total, perm_certified, perm_statuses, perm_entities = summarize_workbook(
        args.perm, "EMP_BUSINESS_NAME", "CASE_STATUS", entity_map
    )

    companies = {}
    for company_id in entity_map:
        companies[company_id] = {
            "lcaCertified": lca_certified[company_id],
            "lcaTotal": lca_total[company_id],
            "lcaStatuses": dict(sorted(lca_statuses[company_id].items())),
            "permCertified": perm_certified[company_id],
            "permTotal": perm_total[company_id],
            "permStatuses": dict(sorted(perm_statuses[company_id].items())),
            "matchedEntities": sorted(lca_entities[company_id] | perm_entities[company_id]),
        }

    snapshot = {
        "fiscalYear": 2025,
        "method": "Exact normalized match against the reviewed legal-entity allowlist.",
        "lcaSource": "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/LCA_Disclosure_Data_FY2025_Q4.xlsx",
        "permSource": "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/PERM_Disclosure_Data_FY2025_Q4.xlsx",
        "companies": companies,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(snapshot, indent=2, sort_keys=True) + "\n")
    args.javascript_output.write_text(
        "window.DOL_FY2025_SNAPSHOT = "
        + json.dumps(snapshot, indent=2, sort_keys=True)
        + ";\n"
    )
    print(f"Wrote {len(companies)} employer summaries to {args.output}")


if __name__ == "__main__":
    main()
